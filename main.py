import sys
import os.path
import random
from datetime import timedelta, date
import xml.etree.ElementTree as et
import openpyxl


def main():
    engineering_team = []
    week = []
    next_monday = str(date.today() + timedelta(days=(7 - date.today().weekday())))
    DAYS_OF_WEEK = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday')

    engineer_root = worker_list_loader()

    engineering_team = schedule_builder(engineer_root, engineering_team)

    week = schedule_randomizer(engineering_team, week)

    workbook = excel_writer(DAYS_OF_WEEK, next_monday, week)

    workbook.save(next_monday + ".xlsx")


def worker_list_loader():
    engineer_tree = et.parse("MESS_list.xml")
    engineer_root = engineer_tree.getroot()
    return engineer_root


def schedule_randomizer(engineering_team, week):
    random.shuffle(engineering_team)
    for i in range(5):
        week.insert(i, engineering_team[:])
        random.shuffle(engineering_team)
    return week


def excel_writer(DAYS_OF_WEEK, next_monday, week):
    workbook = openpyxl.Workbook()
    sheet = workbook["Sheet"]
    sheet.title = next_monday
    workday = []
    row = 1
    column = 1
    for workers_of_the_day, day_of_week in zip(week, DAYS_OF_WEEK):
        sheet.cell(row=row, column=column, value=day_of_week)
        row += 1
        for person in workers_of_the_day:
            workday.append(str(person[1]) + "(" + str(person[0]) + ")" + str(person[2]))
        for i in workday:
            sheet.cell(row=row, column=column, value=i)
            row += 1
        row = 1
        column += 2
        workday = []
    return workbook


def schedule_builder(engineer_root, group):
    for child in engineer_root.findall("Eng"):
        id = child.attrib
        if child[2].text == 'RTP' and child[1].text != 'CP':  # 'CP' is cherry picker, new person
            group.append((id['CEC'], child[0].text, child[1].text, child[2].text))
    return group


if __name__ == '__main__':
    try:
        assert sys.version_info[0] >= 3, "Incorrect interpreter being run. Please use Python 3.x or higher"
        assert os.path.isfile("MESS_list.xml")
    except AssertionError as e:
        print(e)
        exit()
    main()
